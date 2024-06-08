import time
from flask import Flask, render_template, request, jsonify
from excel import append_to_excel, calculate_total_monthly_amount
import openpyxl
from datetime import datetime
from flask_serial import Serial
import serial.tools.list_ports as pyport
import re

def find_arduino_port():
    ports = list(pyport.comports())
    for port in ports:
        if 'Arduino' in port.description:
            return port.device
    return None
arduino_port = find_arduino_port()

app = Flask(__name__)
app.config['SERIAL_TIMEOUT'] = 0.2
app.config['SERIAL_PORT'] = arduino_port
app.config['SERIAL_BAUDRATE'] = 9600
app.config['SERIAL_BYTESIZE'] = 8
app.config['SERIAL_PARITY'] = 'N'
app.config['SERIAL_STOPBITS'] = 1
serial = Serial(app)

DATABASE = 'database.db'
FILE_PATH = 'data.xlsx'
cur_user = None

response_message = None

@serial.on_message()
def handle_message(msg):
    global response_message
    try:
        # print(f"This is msg: {msg}")
        # value = msg.decode('utf-8').replace('\n', '')
        float_value = float(msg.decode())
        # int_value = int(value)
        # print(f"This is int_value: {int_value}")
        if isinstance(float_value, float):
            response_message = float_value
            print(f"This is value:{response_message}")
    # print(response_message)
    except ValueError:
        print("Received message is not a valid float")


# @serial.on_message()
# def handle_message(msg):
#     global response_message
#     text = msg.decode('utf-8').replace('\n', '')
#     response_message = text
#     print(response_message)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/qr_reader')
def qr_reader():
    return render_template('qr_reader.html')

@app.route('/main_menu')
def main_menu():
    name = request.args.get('name')
    phone = request.args.get('phone')
    user_id = request.args.get('id')
    return render_template('main_menu.html', name=name, phone=phone, user_id=user_id)

@app.route('/dispose_oil')
def dispose_oil():
    name = request.args.get('name')
    phone = request.args.get('phone')
    user_id = request.args.get('id')
    return render_template('dispose_oil.html', name=name, phone=phone, user_id=user_id)

# @app.route('/open_lid')
# def open_lid():
#     global response_message
#     response_message = None  # Reset response message
#     try:
#         serial.on_send('O')
        
#         # Wait for response
#         timeout = 5  # seconds
#         start_time = time.time()
#         while response_message is None and (time.time() - start_time) < timeout:
#             time.sleep(0.1)
        
#         if response_message:
#             return jsonify({'status': 'success', 'weight': response_message})
#         else:
#             return jsonify({'status': 'fail', 'error': 'No response within timeout period'})
#     except Exception as e:
#         print(f"Error opening lid: {e}")
#         return jsonify({'status': 'fail', 'error': str(e)})
@app.route('/open_lid')
def open_lid():
    global response_message
    # response_message = None  # Reset response message
    try:
        serial.on_send('O')
        
        # Wait for response
        timeout = 5  # seconds
        start_time = time.time()
        while True:
            # Check if response_message is an float
            if isinstance(response_message, float):
                return jsonify({'status': 'success', 'weight': response_message})
            
            # Check if timeout has occurred
            if (time.time() - start_time) >= timeout:
                return jsonify({'status': 'fail', 'error': 'No response within timeout period'})
            
            time.sleep(0.1)
    except Exception as e:
        print(f"Error opening lid: {e}")
        return jsonify({'status': 'fail', 'error': str(e)})
    

@app.route('/close_lid')
def close_lid():
    global response_message
    # response_message = None  # Reset response message
    try:
        serial.on_send('C')
        
        # Wait for response
        timeout = 20  # seconds
        start_time = time.time()
        # while response_message is None and (time.time() - start_time) < timeout:
        #     time.sleep(0.1)
        if isinstance(response_message, float):
            return jsonify({'status': 'success', 'weight': response_message})
        if (time.time() - start_time) >= timeout:
            return jsonify({'status': 'fail', 'error': 'No response within timeout period'})
        # if response_message:
        #     return jsonify({'status': 'success', 'weight': response_message})
        # else:
        #     return jsonify({'status': 'fail', 'error': 'No response within timeout period'})
    except Exception as e:
        print(f"Error closing lid: {e}")
        return jsonify({'status': 'fail', 'error': str(e)})

@app.route('/calc_weight_again')
def calc_weight_again():
    global response_message
    response_message = None  # Reset response message
    try:
        serial.on_send('R')
        
        # Wait for response
        timeout = 5  # seconds
        start_time = time.time()
        while response_message is None and (time.time() - start_time) < timeout:
            time.sleep(0.1)
        
        if response_message:
            return jsonify({'status': 'success', 'weight': response_message})
        else:
            return jsonify({'status': 'fail', 'error': 'No response within timeout period'})
    except Exception as e:
        print(f"Error calculating weight again: {e}")
        return jsonify({'status': 'fail', 'error': str(e)})
    finally:
        serial.on_send('D')

@app.route('/store_oil')
def store_oil():
    global response_message
    response_message = None  # Reset response message
    try:
        serial.on_send('M')
        
        if response_message:
            return jsonify({'status': 'success'})
        else:
            return jsonify({'status': 'fail'})
    except Exception as e:
        print(f"Error storing oil: {e}")
        return jsonify({'status': 'fail', 'error': str(e)})

@app.route('/start_emptying')
def start_emptying():
    global response_message
    response_message = None  # Reset response message
    try:
        serial.on_send('S')
        
        if response_message:
            return jsonify({'status': 'success'})
        else:
            return jsonify({'status': 'fail'})
    except Exception as e:
        print(f"Error starting emptying process: {e}")
        return jsonify({'status': 'fail', 'error': str(e)})

@app.route('/password')
def password():
    return render_template('password.html')

@app.route('/save_to_excel')
def save_to_excel():
    name = request.args.get('name')
    phone = request.args.get('phone')
    user_id = request.args.get('id')
    oil_amount = request.args.get('weight')
    time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    place = "Unknown"

    append_to_excel(name, phone, user_id, time, place, oil_amount, FILE_PATH)
    current_month = datetime.now().month
    total_amount = calculate_total_monthly_amount(user_id, current_month, FILE_PATH)
    return jsonify(totalAmount=total_amount)

@app.route('/fetch_history')
def fetch_history():
    user_id = request.args.get('id')
    month = int(request.args.get('month'))
    try:
        wb = openpyxl.load_workbook(FILE_PATH)
        sheet = wb.active
        history = []
        total_amount = 0
        weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record_id, name, phone, oil_amount, place, record_time = row
            record_date = datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")
            if record_id == user_id and record_date.month == month:
                weekday_str = weekdays[record_date.weekday()]
                formatted_date = record_date.strftime(f"%Y.%m.%d({weekday_str})")
                history.append({
                    "date": formatted_date,
                    "amount": oil_amount
                })
                total_amount += float(oil_amount)

        return jsonify({"history": history, "total": total_amount})
    except FileNotFoundError:
        return jsonify({"history": [], "total": 0})

@app.route('/history')
def history():
    name = request.args.get('name')
    phone = request.args.get('phone')
    user_id = request.args.get('id')
    return render_template('history.html', name=name, phone=phone, user_id=user_id)

@app.route('/process_qr', methods=['POST'])
def process_qr():
    global cur_user

    data = request.get_json()
    name = data.get('name')
    phone = data.get('phone')
    cur_user = data.get('id')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    db = get_db()
    cursor = db.cursor()
    cursor.execute('INSERT INTO log (name, phone, user_id, timestamp) VALUES (?, ?, ?, ?)', (name, phone, cur_user, timestamp))
    db.commit()
    return jsonify({'status': 'success', 'message': 'User data saved.'})

if __name__ == '__main__':
    app.run(debug=True, use_reloader=False)

# import time
# from flask import Flask, render_template, request, jsonify
# from excel import append_to_excel, calculate_total_monthly_amount
# import openpyxl
# from datetime import datetime
# from flask_serial import Serial
# import serial.tools.list_ports as pyport


# def find_arduino_port():
#     ports = list(pyport.comports())
#     for port in ports:
#         if 'Arduino' in port.description:
#             return port.device
#     return None
# arduino_port = find_arduino_port()

# app = Flask(__name__)
# app.config['SERIAL_TIMEOUT'] = 0.2
# app.config['SERIAL_PORT'] = arduino_port
# app.config['SERIAL_BAUDRATE'] = 9600
# app.config['SERIAL_BYTESIZE'] = 8
# app.config['SERIAL_PARITY'] = 'N'
# app.config['SERIAL_STOPBITS'] = 1
# serial = Serial(app)

# DATABASE = 'database.db'
# FILE_PATH = 'data.xlsx'
# cur_user = None

# @serial.on_message()
# def handle_message(msg):
#     print("serial receive message：", msg)

# @app.route('/')
# def index():
#     return render_template('index.html')

# @app.route('/qr_reader')
# def qr_reader():
#     return render_template('qr_reader.html')

# @app.route('/main_menu')
# def main_menu():
#     name = request.args.get('name')
#     phone = request.args.get('phone')
#     user_id = request.args.get('id')
#     return render_template('main_menu.html', name=name, phone=phone, user_id=user_id)

# @app.route('/dispose_oil')
# def dispose_oil():
#     name = request.args.get('name')
#     phone = request.args.get('phone')
#     user_id = request.args.get('id')
#     return render_template('dispose_oil.html', name=name, phone=phone, user_id=user_id)

# @app.route('/open_lid')
# def open_lid():
#     try:
#         serial.on_send('O')
#         return jsonify({'status': 'success'})
#     except Exception as e:
#         print(f"Error opening lid: {e}")
#         return jsonify({'status': 'fail', 'error': str(e)})

# @app.route('/close_lid')
# def close_lid():
#     try:
#         weight = serial.on_send('C')
#         return jsonify({'status': 'success', 'weight': weight})
#     except Exception as e:
#         print(f"Error closing lid: {e}")
#         return jsonify({'status': 'fail', 'error': str(e)})

# @app.route('/calc_weight_again')
# def calc_weight_again():
#     try:
#         weight = serial.on_send('R')
#         return jsonify({'status': 'success', 'weight': weight})
#     except Exception as e:
#         print(f"Error calculating weight again: {e}")
#         return jsonify({'status': 'fail', 'error': str(e)})
#     finally:
#         serial.on_send('D')

# @app.route('/store_oil')
# def store_oil():
#     try:
#         serial.on_send('M')
#         return jsonify({'status': 'success'})
#     except Exception as e:
#         print(f"Error storing oil: {e}")
#         return jsonify({'status': 'fail', 'error': str(e)})

# @app.route('/start_emptying')
# def start_emptying():
#     try:
#         serial.on_send('S')
#         time.sleep(3)
#         return jsonify({'status': 'success'})
#     except Exception as e:
#         print(f"Error starting emptying process: {e}")
#         return jsonify({'status': 'fail', 'error': str(e)})

# @app.route('/password')
# def password():
#     return render_template('password.html')

# @app.route('/save_to_excel')
# def save_to_excel():
#     name = request.args.get('name')
#     phone = request.args.get('phone')
#     user_id = request.args.get('id')
#     oil_amount = request.args.get('amount')
#     time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#     place = "Unknown"

#     append_to_excel(name, phone, user_id, time, place, oil_amount, FILE_PATH)
#     current_month = datetime.now().month
#     total_amount = calculate_total_monthly_amount(user_id, current_month, FILE_PATH)
#     return jsonify(totalAmount=total_amount)

# @app.route('/fetch_history')
# def fetch_history():
#     user_id = request.args.get('id')
#     month = int(request.args.get('month'))
#     try:
#         wb = openpyxl.load_workbook(FILE_PATH)
#         sheet = wb.active
#         history = []
#         total_amount = 0
#         weekdays = ["월", "화", "수", "목", "금", "토", "일"]
        
#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             record_id, name, phone, oil_amount, place, record_time = row
#             record_date = datetime.strptime(record_time, "%Y-%m-%d %H:%M:%S")
#             if record_id == user_id and record_date.month == month:
#                 weekday_str = weekdays[record_date.weekday()]
#                 formatted_date = record_date.strftime(f"%Y.%m.%d({weekday_str})")
#                 history.append({
#                     "date": formatted_date,
#                     "amount": oil_amount
#                 })
#                 total_amount += float(oil_amount)

#         return jsonify({"history": history, "total": total_amount})
#     except FileNotFoundError:
#         return jsonify({"history": [], "total": 0})

# @app.route('/history')
# def history():
#     name = request.args.get('name')
#     phone = request.args.get('phone')
#     user_id = request.args.get('id')
#     return render_template('history.html', name=name, phone=phone, user_id=user_id)

# @app.route('/process_qr', methods=['POST'])
# def process_qr():
#     global cur_user

#     data = request.get_json()
#     name = data.get('name')
#     phone = data.get('phone')
#     cur_user = data.get('id')
#     timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

#     db = get_db()
#     cursor = db.cursor()
#     cursor.execute('INSERT INTO log (name, phone, user_id, timestamp) VALUES (?, ?, ?, ?)', (name, phone, cur_user, timestamp))
#     db.commit()
#     return jsonify({'status': 'success', 'message': 'User data saved.'})

# if __name__ == '__main__':
#     app.run(debug=True, use_reloader=False)
