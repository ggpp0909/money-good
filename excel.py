import openpyxl
from datetime import datetime

def append_to_excel(name, phone, user_id, time, place, oil_amount, file_path):
    # 엑셀 파일 열기 (없으면 새로 생성)
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["User ID", "Name", "Phone Number", "Oil Quantity", "Place", "Time"])
    else:
        sheet = wb.active

    # 정보 추가
    sheet.append([user_id, name, phone, oil_amount, place, time])

    # 엑셀 파일 저장
    wb.save(file_path)
    print(f"Record appended successfully to {file_path}!")

def calculate_total_monthly_amount(customer_ID, month, file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        total = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            user_id, name, phone, oil_amount, place, time = row
            if user_id == customer_ID and datetime.strptime(time, "%Y-%m-%d %H:%M:%S").month == month:
                if oil_amount is not None:
                    total += float(oil_amount)
        return total
    except FileNotFoundError:
        return 0